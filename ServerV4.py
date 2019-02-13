import socket
import sys
import threading
from threading import Timer
import time
import openpyxl
import os
import random
import datetime


class Server:
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    isSetReserveList = False
    serverDay = 0
    serverMonth = 0
    isFirstDateSet = False

    def __init__(self):
        self.sock.bind(('0.0.0.0', 4969))

    def clientHandlerProcess(self,con,a):
        clientHandler(con,a)

    def calculatorHandlerProcess(self,con):
        Calculator(con)

    def TowerHandlerProcess(self,con):
        towerHandler(con)

    def adminHandlerProcess(self,con):
        adminHandler(con)

    def run(self):

        while True:
            self.sock.listen(20)
            con, a = self.sock.accept()
            try:
                connectMsg = con.recv(1024)
            except ConnectionResetError:
                con.close()
                continue
            connectMsg = str(connectMsg,'utf-8')
            identity = connectMsg

            if identity == 'Calculator':
                print("정산기에서 접속하였습니다.")
                aThread = threading.Thread(target = self.calculatorHandlerProcess, args = (con,))
                aThread.daemon = True
                aThread.start()

            elif identity == 'Client':
                print("IP : " + str(a[0]) + " , 예약자가 접속하였습니다.")
                cThread = threading.Thread(target = self.clientHandlerProcess, args =(con,a,))
                cThread.daemon = True
                cThread.start()
                print()

            elif identity == 'parkTower':
                print("주차타워에서 접속하였습니다.")
                pThread = threading.Thread(target = self.TowerHandlerProcess, args = (con,))
                pThread.daemon = True
                pThread.start()
                print()

            elif identity == 'Admin':
                print("관리자가 접속하였습니다.")
                cThread = threading.Thread(target = self.adminHandlerProcess, args =(con,))
                cThread.daemon = True
                cThread.start()
                print()

            else:
                print("허가되지 않은 접근이 있습니다.")
                con.close()

class Calculator: ###비예약 정산도 엑셀에 날짜별로 기록을 다 남겨놔야 함.
    sortMark = '@'
    def __init__(self,con):
        while True:
            try:
                calMsg = con.recv(1024)
            except ConnectionResetError:
                print("정산기에서 접속을 종료하였습니다.")
                con.close()
                break

            calMsg = str(calMsg,'utf-8')
            calMsg = calMsg.split(self.sortMark)
            MsgHeader = calMsg[0]


            ##### 입차하는 데이터 , 여기서 컨디션 값에 o 넣어줘야함. 여기서 컨디션 엑셀처리
            if MsgHeader == 'carIn':
                lotNum = calMsg[1]
                Date_a = self.today_date()
                self.addNonResParking(Date_a,lotNum)


                ## 여기서는 비예약에서 계산된 값 매출엑셀에 더하기
            if MsgHeader == 'parkingFee':
                lotNum = calMsg[1]
                fee = calMsg[2]
###parking_data 가 정산기 지역변수로 있어서 먼저 동작을 하고 저장이 된 후에야 관리자에서 읽을 수 있음.
                Date_a = self.today_date()
                self.addNonResFee(Date_a,fee)
                ##24시간 넘어서 하루지나가면 기록 다 저장되어있진 않음.(할 수 있으면 수정 필요)
                self.carOut(Date_a,lotNum)

                ###보내주는건 관리자 비예약에서 하면되고 , 여기서는 엑셀에 저장만 하기

#############함수처리
    def __call__(self,*args,**kwargs):
        print("정산기에서 접속을 종료하였습니다.")

    def today_date(self):

        d = datetime.date.today()


        month_a = str(d.month)
        date_a = str(d.day)
        ##밑에는 인트값임 쓸거면 그때 리턴 처리 추가
        month_d = d.month
        date_d = d.day

        Date_a = month_a + "." + date_a

        return Date_a


        ############정산기 동작하면서 입차기록
    def addNonResParking(self,today_date,lotNum):

        exDoc_4 = openpyxl.load_workbook('NonResParking.xlsx')


        if today_date not in exDoc_4.sheetnames:       #입력받은 날짜의 시트가 없을 경우 셋팅
            exDoc_4.save('NonResParking.xlsx')
            exDoc_4 = self.setNewNonResDay(today_date)

        sheet_4 = exDoc_4[today_date]

        lotNum = int(lotNum)

        sheet_4.cell(row = 2, column = lotNum).value = 'o'

        exDoc_4.save('NonResParking.xlsx')
        exDoc_4.close()

    def setNewNonResDay(self,today_date):
        exDoc = openpyxl.load_workbook('NonResParking.xlsx')


        newDay = exDoc.create_sheet(today_date)
        for row in newDay.iter_rows(min_row=1, max_row=2, min_col=1, max_col=20):
            for cell in row:
                if cell.row == 1:
                    cell.value = '주차장자리' + str(cell.column)
                else:
                    cell.value = 'x'
        exDoc.save('NonResParking.xlsx')
        exDoc = openpyxl.load_workbook('NonResParking.xlsx')
        return exDoc




        ####출차하면서 비용 엑셀에 기록
    def addNonResFee(self,Date_a,fee):

        exDoc_4 = openpyxl.load_workbook('earning.xlsx')

        month_date = Date_a.split('.')
        month_a = month_date[0]
        date = month_date[1]
        date = int(date)
        fee = int(fee)


        sheet_4 = exDoc_4[month_a]
        if month_a not in exDoc_4.sheetnames:       #입력받은 날짜의 시트가 없을 경우 셋팅
            exDoc_4.save('earning.xlsx')
            exDoc_4 = self.setNewEarning(Date_a)

        NonResDate_fee = sheet_4.cell(row = 7, column = date).value
        NonResDate_fee += fee
        sheet_4.cell(row = 7, column = date).value = NonResDate_fee

        sum = 0

        for row in sheet_4.iter_rows(min_row=3, max_col=31, max_row=3):
            for cell in row:

                sum += int(cell.value)

        for row in sheet_4.iter_rows(min_row=7, max_col=31, max_row=7):
            for cell in row:

                sum += int(cell.value)


        sheet_4.cell(row = 5, column = 1).value = sum

        exDoc_4.save('earning.xlsx')
        exDoc_4.close()

    def carOut(self,Date_a,lotNum):

        exDoc_4 = openpyxl.load_workbook('NonResParking.xlsx')

        sheet_4 = exDoc_4[Date_a]
        if Date_a not in exDoc_4.sheetnames:       #입력받은 날짜의 시트가 없을 경우 셋팅
            exDoc_4.save('NonResParking.xlsx')
            exDoc_4 = self.setNewNonResDay(Date_a)

        lotNum = int(lotNum)
        print(lotNum)
        sheet_4.cell(row = 2, column = lotNum).value = 'x'

        exDoc_4.save('NonResParking.xlsx')
        exDoc_4.close()

    def setNewEarning(self,resDate):
        date_number = 1
        exDoc = openpyxl.load_workbook('earning.xlsx')
        month_date = resDate.split('.')
        sheet = month_date[0]
        month_earning = exDoc.create_sheet(sheet)
        #글자 조정하기
        month_earning.merge_cells('A1:AE1')
        month_earning['A1'] = '일별매출관리'
        day_earning_manager = month_earning['A1']
        month_earning.merge_cells('A4:C4')
        month_earning['A4'] = '이번 달 매출관리'
        month_earning_manager = month_earning['A1']
        month_earning.merge_cells('A6:AE6')
        month_earning['A1'] = '비예약일별매출관리'

        for row in month_earning.iter_rows(min_row=2, max_row=7, min_col=1, max_col=31):
            for cell in row:
                if cell.row == 2:
                    cell.value = sheet + "." + (str(date_number))
                    date_number += 1
                if cell.row == 3:
                    cell.value = 0
                    ##비예약
                if cell.row == 7:
                    cell.value = 0


class clientHandler:
    sortMark = '@'
    def __init__(self,con,a):
        while True:
            try:
                clientMsg = con.recv(1024)
            except ConnectionResetError:
                print("IP : " + a[0] + " 접속자가 접속을 해제하였습니다.")
                con.close()
                break

            clientMsg = str(clientMsg,'utf-8')
            clientMsg = clientMsg.split(self.sortMark)
            MsgHeader = clientMsg[0]

            if MsgHeader == 'reqResList':
                reqDate = clientMsg[1] #11.21
                parkLot1 = self.loadParkLot(reqDate,1) #'1 2 3  5 6 7 8 9 ..'
                parkLot2 = self.loadParkLot(reqDate,2)
                parkLot3 = self.loadParkLot(reqDate,3)
                parkLot4 = self.loadParkLot(reqDate,4)
                ResList = parkLot1 + self.sortMark + parkLot2 + self.sortMark + parkLot3 + self.sortMark + parkLot4
                serverHeader = 'ResList'
                serverMsg = serverHeader + self.sortMark + ResList    #ResList@1 2 3..@2 4 5 6..@.....
                con.send(bytes(serverMsg,'utf-8'))

            elif MsgHeader == 'postRes':
                resDate = clientMsg[1]      #'12.25'
                resParkLot = clientMsg[2]   #1
                resStart = clientMsg[3]     #17
                resEnd = clientMsg[4]       #19
                self.makeRes(resDate,resParkLot,resStart,resEnd)
                ACode = self.saveACode(resDate,resParkLot,resStart,resEnd)
                serverHeader = 'ResSuc'
                serverMsg = serverHeader + self.sortMark + ACode
                con.send(bytes(serverMsg,'utf-8'))      #postRes@6875

            elif MsgHeader == 'inqResTime':
                clientACode = clientMsg[1]
                resDate, resTime, resLot =  self.inqTime(clientACode)   #'11.22','3,4,5','3'
                serverHeader = 'inqTime'
                if resDate == None:
                    serverMsg = serverHeader + self.sortMark + str(resDate)
                else:
                    serverMsg = serverHeader + self.sortMark + resDate + self.sortMark + resTime + self.sortMark + resLot
                con.send(bytes(serverMsg,'utf-8'))

    def __call__(self,*args,**kwargs):
        print("종료합니다.")

    def inqTime(self,clientACode):
        exDoc = openpyxl.load_workbook('ACode.xlsx')
        sheetList = []
        resDate = []
        resTime = ''
        resLot = []
        for i in exDoc.sheetnames:
            sheetList.append(i)
        for j in sheetList:
            reqSheet = exDoc[j]
            for row in reqSheet.iter_rows(min_row=2, max_row=25, min_col=1, max_col=4):
                for cell in row:
                    if str(cell.value) == clientACode:
                        resDate.append(j)
                        resTime += str(cell.row-2) + ','
                        resLot.append(cell.column)
        if len(resDate) == 0:
            return (None,None,None)
        else:
            return (resDate[0],resTime,resLot[0][-1])

    def makeRes(self,resDate,resParkLot,resStart,resEnd):
        exDoc = openpyxl.load_workbook('ReserveList.xlsx')
        reqSheet = exDoc[resDate]
        for i in range(int(resStart),int(resEnd)):
            reqSheet.cell(row = i+2, column = int(resParkLot)).value = None
        exDoc.save('ReserveList.xlsx')

    def saveACode(self,resDate,resParkLot,resStart,resEnd):
        ACode = ''.join(["%s" % random.randint(0, 9) for num in range(0, 4)])
        exDoc = openpyxl.load_workbook('ACode.xlsx')
        reqSheet = exDoc[resDate]
        for i in range(int(resStart),int(resEnd)):
            reqSheet.cell(row = i+2, column = int(resParkLot)).value = ACode
        exDoc.save('ACode.xlsx')
        return ACode

    def setNewDay(self,sheet):
        exDoc = openpyxl.load_workbook('ReserveList.xlsx')
        newDay = exDoc.create_sheet(sheet)
        for row in newDay.iter_rows(min_row=1, max_row=25, min_col=1, max_col=4):
            for cell in row:
                if cell.row == 1:
                    cell.value = 'parkLot' + str(cell.column)
                else:
                    cell.value = str(cell.row-2)
        exDoc.save('ReserveList.xlsx')
        exDoc = openpyxl.load_workbook('ReserveList.xlsx')
        return exDoc

    def setNewACode(self,sheet):
        exDoc = openpyxl.load_workbook('ACode.xlsx')
        newACode = exDoc.create_sheet(sheet)
        for row in newACode.iter_rows(min_row=1, max_row=25, min_col=1, max_col=4):
            for cell in row:
                if cell.row == 1:
                    cell.value = 'parkLot' + str(cell.column)
        exDoc.save('ACode.xlsx')

    def loadParkLot(self,sheet,lotNum):
        exDoc = openpyxl.load_workbook('ReserveList.xlsx')
        if sheet not in exDoc.sheetnames:       #입력받은 날짜의 시트가 없을 경우 셋팅
            exDoc.save('ReserveList.xlsx')
            exDoc = self.setNewDay(sheet)
            self.setNewACode(sheet)
        reqSheet = exDoc[sheet]
        res = ''
        cnt = 0
        for row in reqSheet.iter_rows(min_row=2, max_row=25, min_col=lotNum, max_col=lotNum):
            for cell in row:
                if cnt >= 10:
                    res += str(cell.value).replace("None","  ") + ' '
                else:
                    res += str(cell.value).replace("None"," ") + ' '
                cnt += 1
        return res


class towerHandler:
    sortMark = '@'
    def __init__(self,con):
        while True:
            try:
                towerMsg = con.recv(1024)
            except:
                print("주차타워 접속을 종료합니다.")
                con.close()
                break

            towerMsg = str(towerMsg,'utf-8')
            towerMsg = towerMsg.split(self.sortMark) #['checkACode']
            MsgHeader = towerMsg[0]

            if MsgHeader == 'checkACode':        #checkACode@키패드입력인증코드
                rcvACode = towerMsg[1]
                isPass , reqParkLot = self.checkACode(rcvACode)
                serverHeader = 'ACode'
                reqParkLot = str(reqParkLot)
                serverMsg = serverHeader + self.sortMark + isPass + self.sortMark + reqParkLot
                con.send(bytes(serverMsg,'utf-8'))      #checkACode@True@3 or checkACode@False@None
            else:
                pass

    def checkACode(self,rcvACode):                     #인증코드 요청한 시간대의 인증코드 반환
        curMonth = str(datetime.datetime.today().month)
        curDay = str(datetime.datetime.today().day)
        reqDate = curMonth + '.' + curDay               #엑셀 시트
        curTime = int(time.ctime().split(' ')[3][:2])   #인증코드를 불러와야 할 셀
        exDoc = openpyxl.load_workbook('ACode.xlsx')
        reqSheet = exDoc[reqDate]
        for c in range(1,5):
            ACode = reqSheet.cell(row=curTime+2,column=c).value
            if rcvACode == ACode:
                reqParkLot = reqSheet.cell(row=1,column=c).value
                return (str(True),reqParkLot[-1])
        return (str(False),None)

    def __call__(self,*args,**kwargs):
        print("주차타워 접속을 종료합니다.")

class adminHandler:
    sortMark = '@'

    NonResData = None

    def __init__(self,con):
        while True:
            try:
                adminMsg = con.recv(1024)
            except ConnectionResetError:
                print("관리자 접속을 종료합니다.")
                con.close()
                break

            adminMsg = str(adminMsg,'utf-8')
            adminMsg = adminMsg.split(self.sortMark)
            MsgHeader = adminMsg[0]

######      비예약
            if MsgHeader == 'reqNonResList':   #리스트 띄우게 보내달라고 요청하는 것
                #딕션으로 받음


                Date_a = self.today_date()
                NonResParkingLot = self.NonResParkingList(Date_a)
                NonResList = NonResParkingLot
                serverHeader = 'NonResList'
                serverMsg = serverHeader + self.sortMark + NonResList    #ResList@1 2 3..@2 4 5 6..@.....
                con.send(bytes(serverMsg,'utf-8'))


#############예약
            if MsgHeader == 'reqResList':   #리스트 띄우게 보내달라고 요청하는 것
                reqDate = adminMsg[1] #11.21
                parkLot1 = self.loadParkLot(reqDate,1) #'1 2 3  5 6 7 8 9 ..'
                parkLot2 = self.loadParkLot(reqDate,2)
                parkLot3 = self.loadParkLot(reqDate,3)
                parkLot4 = self.loadParkLot(reqDate,4)
                ResList = parkLot1 + self.sortMark + parkLot2 + self.sortMark + parkLot3 + self.sortMark + parkLot4
                serverHeader = 'ResList'
                serverMsg = serverHeader + self.sortMark + ResList    #ResList@1 2 3..@2 4 5 6..@.....
                con.send(bytes(serverMsg,'utf-8'))

            elif MsgHeader == 'postRes':   #예약을 넣는 요청 (날짜 장소 예약시간 파라미터로 날려서 makeRes 함수 실행시켜서 안에 값 비우게 하는 것)
                resDate = adminMsg[1]      #'12.25'
                resParkLot = adminMsg[2]   #1
                resStart = adminMsg[3]     #17
                resEnd = adminMsg[4]       #19
                self.makeRes(resDate,resParkLot,resStart,resEnd)
                date_earning = self.makeRsvEarning(resDate,resParkLot,resStart,resEnd)  ##매출 추가 하는것
                self.addEarning(resDate,date_earning)
                ACode = self.saveACode(resDate,resParkLot,resStart,resEnd)    ##파라미터에 날짜 자리 예약시간보내서 인증코드 보내는 것
                serverHeader = 'ResSuc'
                serverMsg = serverHeader + self.sortMark + ACode    #ResSuc@5858이런식으로 성공했다고 날려주는 작업임.
                con.send(bytes(serverMsg,'utf-8'))      #postRes@6875

            elif MsgHeader == 'inqResTime':
                adminACode = adminMsg[1]
                resDate, resTime, resLot =  self.inqTime(adminACode)   #'11.22','3,4,5','3'
                serverHeader = 'inqTime'
                if resDate == None:
                    serverMsg = serverHeader + self.sortMark + str(resDate)
                else:
                    serverMsg = serverHeader + self.sortMark + resDate + self.sortMark + resTime + self.sortMark + resLot
                con.send(bytes(serverMsg,'utf-8'))

                ##여기서부터 예약 삭제기능

            elif MsgHeader == 'delRes':   #삭제 요청 받는거
                resDate = adminMsg[1]
                resParkLot = adminMsg[2]
                resStart = adminMsg[3]
                resEnd = adminMsg[4]
                self.delRes(resDate,resParkLot,resStart,resEnd)
                date_earning = self.makeRsvEarning(resDate,resParkLot,resStart,resEnd)  ##매출 추가 하는것
                self.addEarning(resDate,date_earning)


###여기서부터 매출리스트 기능

            elif MsgHeader == 'reqEarningList':
                sum = 0
                resDate = adminMsg[1]
                earningList_month, earningList_date = self.earningList(resDate)
                earningList = str(earningList_month) + "@" + str(earningList_date)
                EarningList = str(earningList)
                serverHeader = 'EarningList'
                serverMsg = serverHeader + self.sortMark + EarningList    #ResList@1 2 3..@2 4 5 6..@.....
                con.send(bytes(serverMsg,'utf-8'))

            elif MsgHeader == 'reqEarningList_month':
                sum = 0
                seeingMonth = adminMsg[1]
                earningList_month = self.earningList_month(seeingMonth)
                earningList_month = str(earningList_month)
                EarningList_month = str(earningList_month)
                serverHeader = 'EarningList_month'
                serverMsg = serverHeader + self.sortMark + EarningList_month   #ResList@1 2 3..@2 4 5 6..@.....
                con.send(bytes(serverMsg,'utf-8'))





###############함수 정의 ###############################

####비예약
    def today_date(self):

        d = datetime.date.today()


        month_a = str(d.month)
        date_a = str(d.day)
        ##밑에는 인트값임 쓸거면 그때 리턴 처리 추가
        month_d = d.month
        date_d = d.day

        Date_a = month_a + "." + date_a

        return Date_a

    def NonResParkingList(self,Date_a):

        exDoc_4 = openpyxl.load_workbook('NonResParking.xlsx')


        sheet_3 = exDoc_4[Date_a]

        NonResParkingLot = ''
        for row in sheet_3.iter_rows(min_row=2, max_col=20, max_row=2):
            for cell in row:


                NonResParkingLot += cell.value

        if Date_a not in exDoc_4.sheetnames: ##없으면 없다고 말하고 , 관리자에서 컨티뉴로 돌아가기
            print(Date_a,"달 비예약 주차장 주차상태가 없습니다.")


        return NonResParkingLot




############예약
    def inqTime(self,adminACode):
        exDoc = openpyxl.load_workbook('ACode.xlsx')
        sheetList = []
        resDate = []
        resTime = ''
        resLot = []
        for i in exDoc.sheetnames:
            sheetList.append(i)
        for j in sheetList:
            reqSheet = exDoc[j]
            for row in reqSheet.iter_rows(min_row=2, max_row=25, min_col=1, max_col=4):
                for cell in row:
                    if str(cell.value) == adminACode:
                        resDate.append(j)
                        resTime += str(cell.row-2) + ','
                        resLot.append(cell.column)
        if len(resDate) == 0:
            return (None,None,None)
        else:
            return (resDate[0],resTime,resLot[0][-1])

    def makeRes(self,resDate,resParkLot,resStart,resEnd):   #예약 만드는 작업 리저브리스트 엑셀 열어서, 안에다가 숫자 빼는 작업
        exDoc = openpyxl.load_workbook('ReserveList.xlsx')
        reqSheet = exDoc[resDate]
        for i in range(int(resStart),int(resEnd)):
            reqSheet.cell(row = i+2, column = int(resParkLot)).value = None
        exDoc.save('ReserveList.xlsx')

    def saveACode(self,resDate,resParkLot,resStart,resEnd):  #예약 인증코드 발급해서 엑셀 안에다가 넣는 작업임.
        ACode = ''.join(["%s" % random.randint(0, 9) for num in range(0, 4)])
        exDoc = openpyxl.load_workbook('ACode.xlsx')
        reqSheet = exDoc[resDate]
        for i in range(int(resStart),int(resEnd)):
            reqSheet.cell(row = i+2, column = int(resParkLot)).value = ACode
        exDoc.save('ACode.xlsx')
        return ACode

    def setNewDay(self,sheet):    #엑셀에 예약하는 새로운 날 시트 만드는 함수
        exDoc = openpyxl.load_workbook('ReserveList.xlsx')
        newDay = exDoc.create_sheet(sheet)
        for row in newDay.iter_rows(min_row=1, max_row=25, min_col=1, max_col=4):
            for cell in row:
                if cell.row == 1:
                    cell.value = 'parkLot' + str(cell.column)
                else:
                    cell.value = str(cell.row-2)
        exDoc.save('ReserveList.xlsx')
        exDoc = openpyxl.load_workbook('ReserveList.xlsx')
        return exDoc

    def setNewACode(self,sheet):
        exDoc = openpyxl.load_workbook('ACode.xlsx')
        newACode = exDoc.create_sheet(sheet)
        for row in newACode.iter_rows(min_row=1, max_row=25, min_col=1, max_col=4):
            for cell in row:
                if cell.row == 1:
                    cell.value = 'parkLot' + str(cell.column)
        exDoc.save('ACode.xlsx')

    def loadParkLot(self,sheet,lotNum):
        exDoc = openpyxl.load_workbook('ReserveList.xlsx')
        if sheet not in exDoc.sheetnames:       #입력받은 날짜의 시트가 없을 경우 셋팅
            exDoc.save('ReserveList.xlsx')
            exDoc = self.setNewDay(sheet)
            self.setNewACode(sheet)

    #    exDoc_4 = openpyxl.load_workbook('earning.xlsx')

    #    month_date = sheet.split('.')
    #    month_a = month_date[0]
    #    date = month_date[1]
    #    date = int(date)
    #    sheet_4 = exDoc_4[month_a]
#
#        if month_a not in exDoc_4.sheetnames:       #입력받은 날짜의 시트가 없을 경우 셋팅
#            exDoc_4.save('earning.xlsx')
#            exDoc_4 = self.setNewEarning(sheet)

        reqSheet = exDoc[sheet]
        res = ''
        cnt = 0
        for row in reqSheet.iter_rows(min_row=2, max_row=25, min_col=lotNum, max_col=lotNum):
            for cell in row:
                if cnt >= 11:
                    res += str(cell.value).replace("None","  ") + ' '
                else:
                    res += str(cell.value).replace("None"," ") + ' '
                cnt += 1
        return res



#삭제 기능 (여기 안에는 예약 엑셀파일 안에 숫자 다시 추가하고 , 인증코드 엑셀에는 인증코드 삭제하는 기능이 들어가야함)
    def delRes(self,resDate,resParkLot,resStart,resEnd):
        #def makeRes(self,resDate,resParkLot,resStart,resEnd):
        exDoc_1 = openpyxl.load_workbook('ReserveList.xlsx')
        reqSheet_1 = exDoc_1[resDate]
        for i in range(int(resStart),int(resEnd)):
            reqSheet_1.cell(row = i+2, column = int(resParkLot)).value = str(i) #여기에 원래 있던 숫자값 추가하기
        exDoc_1.save('ReserveList.xlsx')
        exDoc_1.close()
        # def saveACode(self,resDate,resParkLot,resStart,resEnd):  #예약 인증코드 발급해서 엑셀 안에다가 넣는 작업임.
        # ACode = ''.join(["%s" % random.randint(0, 9) for num in range(0, 4)])
        exDoc_2 = openpyxl.load_workbook('ACode.xlsx')
        reqSheet_2 = exDoc_2[resDate]
        for i in range(int(resStart),int(resEnd)):
            reqSheet_2.cell(row = i+2, column = int(resParkLot)).value = None
        exDoc_2.save('ACode.xlsx')
        exDoc_2.close()


#######여기는 매출관리 함수

###이거는 매출추가 함수임 이건 예약 발생하거나 삭제 발생할 때마다 뒤에서 실행시켜 줘야함
    def makeRsvEarning(self,resDate,resParkLot,resStart,resEnd):
        exDoc_1 = openpyxl.load_workbook('ReserveList.xlsx')
        reqSheet_1 = exDoc_1[resDate]
        count = 0
        for row in reqSheet_1.iter_rows(min_row=2,  max_col=4 , max_row=25):

            for cell in row:

                if cell.value == None:

                    count += 1
        #count 는 예약 개수
        date_earning = count * 2000
        #예약 시간당 1000원
        return date_earning

    def addEarning(self,resDate,date_earning):
        ###여기는 earning 엑셀 시트마다 달 별로 분류되어있음 . 위에서 뽑아온거 기록하면
        exDoc_4 = openpyxl.load_workbook('earning.xlsx')

        month_date = resDate.split('.')
        month_a = month_date[0]
        date = month_date[1]
        date = int(date)

        if month_a not in exDoc_4.sheetnames:       #입력받은 날짜의 시트가 없을 경우 셋팅
            exDoc_4.save('earning.xlsx')
            exDoc_4 = self.setNewEarning(resDate)

        sheet_4 = exDoc_4[month_a]

        sheet_4.cell(row = 3, column = date).value = date_earning

        sum = 0

        for row in sheet_4.iter_rows(min_row=3, max_col=31, max_row=3):
            for cell in row:

                sum += int(cell.value)

        for row in sheet_4.iter_rows(min_row=7, max_col=31, max_row=7):
            for cell in row:

                sum += int(cell.value)

        sheet_4.cell(row = 5, column = 1).value = sum

        exDoc_4.save('earning.xlsx')
        exDoc_4.close()

    def setNewEarning(self,resDate):
        date_number = 1
        exDoc_3 = openpyxl.load_workbook('earning.xlsx')
        month_date = resDate.split('.')
        sheet = month_date[0]
        month_earning = exDoc_3.create_sheet(sheet)
        #글자 조정하기
        month_earning.merge_cells('A1:AE1')
        month_earning['A1'] = '일별매출관리'
        day_earning_manager = month_earning['A1']
        month_earning.merge_cells('A4:C4')
        month_earning['A4'] = '이번 달 매출관리'
        month_earning_manager = month_earning['A1']
        month_earning.merge_cells('A6:AE6')
        month_earning['A6'] = '비예약일별매출관리'

        for row in month_earning.iter_rows(min_row=2, max_row=7, min_col=1, max_col=31):
            for cell in row:
                if cell.row == 2:
                    cell.value = sheet + "." + (str(date_number))
                    date_number += 1
                if cell.row == 3:
                    cell.value = 0
                    ##비예약
                if cell.row == 7:
                    cell.value = 0


        exDoc_3.save('earning.xlsx')
        exDoc = openpyxl.load_workbook('earning.xlsx')
        return exDoc


    def earningList(self, resDate):  ##리스트 띄워주는거 구현해야 됨
        exDoc_4 = openpyxl.load_workbook('earning.xlsx')
        month_date = resDate.split('.')
        sheet = month_date[0]

        date = month_date[1]
        date = int(date)

        sheet_3 = exDoc_4[sheet]

        earningList_month = None
        earningList_date = None


        earningList_month = sheet_3.cell(row = 5, column = 1).value
        earningList_date = sheet_3.cell(row = 3, column = date).value
        earningList_date += sheet_3.cell(row = 7, column = date).value
        ###이것도 출력하는거 만들어야됨 아직 테스트 안됨
        ## earningList_date_nonres = sheet_3.cell(row = 7, column = date).value



        if sheet not in exDoc_4.sheetnames: ##없으면 없다고 말하고 , 관리자에서 컨티뉴로 돌아가기
            print(sheet,"달에 기록된 매출이 없습니다.")


        return earningList_month , earningList_date


    def earningList_month(self, seeingMonth):  ##리스트 띄워주는거 구현해야 됨
        exDoc_4 = openpyxl.load_workbook('earning.xlsx')

        sheet = seeingMonth
        sheet_3 = exDoc_4[sheet]
        earningList_month = None
        earningList_month = sheet_3.cell(row = 5, column = 1).value

        if sheet not in exDoc_4.sheetnames: ##없으면 없다고 말하고 , 관리자에서 컨티뉴로 돌아가기
            print(sheet,"달에 기록된 매출이 없습니다.")


        return earningList_month





server = Server()
server.run()
